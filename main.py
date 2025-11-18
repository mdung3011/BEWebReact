from flask import Flask, request, jsonify,send_file
from flask_cors import CORS
from db import get_connection
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from datetime import datetime
import calendar
app = Flask(__name__)
CORS(app)
def get_days_in_month(month: int) -> int:
    """Trả về số ngày trong tháng (không phân biệt năm, Feb = 28)."""
    if month in (1, 3, 5, 7, 8, 10, 12):
        return 31
    if month in (4, 6, 9, 11):
        return 30
    return 28  # tháng 2
nam = datetime.now().year

def format_rows(rows):
    def fmt_date(d):
        return d.strftime("%Y-%m-%d") if d else None

    def fmt_dt(dt):
        return dt.strftime("%Y-%m-%dT%H:%M") if dt else None

    result = []
    for (line, machine, day, dayPlan, target, cycle, s1, e1, s2, e2, pid) in rows:
        result.append({
            "line": line,
            "machine": machine,
            "day": fmt_date(day),
            "dayPlan": dayPlan,
            "targetProduct": target,
            "cycleTime": cycle,
            "startShift1": fmt_dt(s1),
            "endShift1": fmt_dt(e1),
            "startShift2": fmt_dt(s2),
            "endShift2": fmt_dt(e2),
            "id": pid
        })
    return result
@app.route("/api/login", methods=["POST"])
def login():
    data = request.get_json()
    username = data.get("username")
    password = data.get("password")

    if not username or not password:
        return jsonify({"ok": False, "message": "Thiếu tài khoản hoặc mật khẩu"}), 400

    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM users WHERE username=%s AND password=%s", (username, password))
    user = cur.fetchone()
    cur.close()
    conn.close()

    if user:
        return jsonify({"ok": True, "user": {"username": user["username"], "full_name": user["full_name"]}})
    else:
        return jsonify({"ok": False, "message": "Sai tài khoản hoặc mật khẩu"}), 401
@app.route("/api/register", methods=["POST"])
def register():
    data = request.get_json(force=True) or {}
    print("DEBUG /api/register payload:", data)  # xem payload thực tế

    def as_text(v):
        if isinstance(v, str):
            return v
        if isinstance(v, (int, float)):
            return str(v)
        if isinstance(v, dict):
            # cố gắng lấy các key hay gặp khi gửi nhầm object
            for k in ("value", "username", "name"):
                if isinstance(v.get(k), str):
                    return v[k]
        return ""

    username = as_text(data.get("username")).strip()
    password = as_text(data.get("password")).strip()
    full_name = as_text(data.get("full_name")).strip()

    if not username or not password or not full_name:
        return jsonify({"ok": False, "message": "Thiếu username/password/full_name"}), 400

    conn = get_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute("SELECT id FROM users WHERE username=%s LIMIT 1", (username,))
    if cur.fetchone():
        cur.close();
        conn.close()
        return jsonify({"ok": False, "message": "Tài khoản đã tồn tại"}), 409

    cur.execute(
        "INSERT INTO users (username, password, full_name) VALUES (%s, %s, %s)",
        (username, password, full_name)
    )
    conn.commit()
    cur.close();
    conn.close()
    return jsonify({"ok": True, "message": "Đăng ký thành công"})
@app.route("/api/lines")
def get_lines():
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT 
            LineID   AS idline,
            LineName AS ten_line
        FROM productionline
    """)

    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    return jsonify(rows)
@app.route("/api/lines/<int:idline>/machines")
def get_machines_by_line(idline):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT 
            MachineID   AS id,
            MachineName AS name
        FROM machine
        WHERE LineID = %s
    """, (idline,))

    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(rows)
@app.route("/api/machines/<int:machine_id>/day")
def get_machine_day(machine_id):
    day = request.args.get("day")
    if not day:
        return jsonify({"error": "Missing day param"}), 400

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    # --------- LẤY DỮ LIỆU THỜI GIAN (dayvalues) ----------
    cursor.execute("""
        SELECT 
            Days,
            PowerRun,
            Operation,
            SmallStop,
            Fault,
            Break,
            Maintenance,
            Eat,
            Waiting,
            MachineryEdit,
            ChangeProductCode,
            Glue_CleaningPaper,
            Others
        FROM dayvalues
        WHERE MachineID = %s AND Days = %s
        LIMIT 1
    """, (machine_id, day))

    row = cursor.fetchone()

    # (tí nữa còn dùng connection, đừng đóng vội)
    if not row:
        cursor.close()
        conn.close()
        return jsonify({
            "machine_id": machine_id,
            "day": day,
            "data": None
        })

    # ---- POWER RUN: 2 chữ sau dấu chấm ----
    raw_power = row.get("PowerRun")
    try:
        power_val = float(raw_power) if raw_power else 0.0
    except:
        power_val = 0.0
    power_run_str = f"{power_val:.2f}"

    # ---- CÁC CATEGORY (cho pie + bảng) ----
    categories_raw = {
        "Operation":          float(row["Operation"]          or 0.0),
        "SmallStop":          float(row["SmallStop"]          or 0.0),
        "Fault":              float(row["Fault"]              or 0.0),
        "Break":              float(row["Break"]              or 0.0),
        "Maintenance":        float(row["Maintenance"]        or 0.0),
        "Eat":                float(row["Eat"]                or 0.0),
        "Waiting":            float(row["Waiting"]            or 0.0),
        "MachineryEdit":      float(row["MachineryEdit"]      or 0.0),
        "ChangeProductCode":  float(row["ChangeProductCode"]  or 0.0),
        "Glue_CleaningPaper": float(row["Glue_CleaningPaper"] or 0.0),
        "Others":             float(row["Others"]             or 0.0),
    }

    total_hours = sum(categories_raw.values())
    if total_hours <= 0:
        total_hours = 1.0

    color_map = {
        "Operation":          "#00a03e",
        "SmallStop":          "#f97316",
        "Fault":              "#ef4444",
        "Break":              "#eab308",
        "Maintenance":        "#6b21a8",
        "Eat":                "#22c55e",
        "Waiting":            "#0ea5e9",
        "MachineryEdit":      "#1d4ed8",
        "ChangeProductCode":  "#a855f7",
        "Glue_CleaningPaper": "#fb7185",
        "Others":             "#6b7280",
    }

    detail_rows = []
    pie_data = []

    for label, value in categories_raw.items():
        hours = float(value)
        h = int(hours)
        m = int(round((hours - h) * 60))
        time_str = f"{h}h {m}m"

        ratio = round((hours / total_hours) * 100.0, 2)
        ratio_text = f"{ratio:.2f}%"

        detail_rows.append({
            "label": label,
            "value": hours,
            "time": time_str,
            "ratio": ratio,
            "ratio_text": ratio_text,
            "color": color_map[label],
        })

        pie_data.append({
            "name": label,
            "value": ratio,
            "color": color_map[label],
        })

    # --------- THÊM PHẦN PRODUCT: TOTAL / OK / NG / RATIO ----------
    # TODO: sửa lại tên bảng + cột cho đúng DB thực tế của bạn
    #
    # Ví dụ: bảng dayproduct có cột:
    #   MachineID, Days, Total, OK, NG
    #
    cursor.execute("""
        SELECT 
            totalproduct_actual AS Total,
            totalproduct_ok as OK,
            totalproduct_ng as NG
        FROM production_output
        WHERE machineid = %s AND days = %s
        LIMIT 1
    """, (machine_id, day))

    prod = cursor.fetchone()
    cursor.close()
    conn.close()

    if prod:
        total = float(prod["Total"] or 0)
        ok = float(prod["OK"] or 0)
        ng = float(prod["NG"] or 0)
    else:
        total, ok, ng = 0.0, 0.0, 0.0

    ratio = (ok * 100.0 / total) if total > 0 else 0.0

    product = {
        "total": int(total),
        "ok": int(ok),
        "ng": int(ng),
        "ratio": round(ratio, 2),
        "ratio_text": f"{ratio:.2f}%"
    }

    # --------- KẾT QUẢ TRẢ VỀ ----------
    result = {
        "machine_id": machine_id,
        "day": row["Days"],
        "power_run": power_run_str,
        "total_hours": round(total_hours, 2),
        "pie": pie_data,
        "details": detail_rows,
        "product": product,          # 👈 FE dùng cho bảng PRODUCT
    }

    return jsonify(result)
from flask import request, jsonify

@app.route("/api/machines/<int:machine_id>/month-ratio")
def get_machine_month_ratio(machine_id):
    try:
        month = int(request.args.get("month"))
    except (TypeError, ValueError):
        return jsonify({"error": "Missing or invalid month param"}), 400

    data_type = request.args.get("data", "")  # VD "OEE RATIO" (để echo lại cho FE)

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    # Lấy TỪNG DÒNG theo ngày trong tháng (không AVG)
    cursor.execute(
        """
        SELECT
            Days,
            OEERatio,
            OKProductRatio,
            OutputRatio,
            ActivityRatio
        FROM sdvn.dayvalues
        WHERE YEAR(Days)= %s AND MachineID = %s
          AND MONTH(Days) = %s
        ORDER BY Days
        """,
        (nam,machine_id, month),
    )

    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    # map theo ngày (1..31) => row
    day_map = {}

    for row in rows:
        day_raw = row["Days"]
        # cố lấy số ngày (1..31)
        if hasattr(day_raw, "day"):
            dnum = day_raw.day
        else:
            # fallback: lấy 2 ký tự cuối, ví dụ "2025-09-05"
            try:
                dnum = int(str(day_raw)[-2:])
            except ValueError:
                continue

        day_map[dnum] = row  # nếu 1 ngày có nhiều dòng, dòng cuối sẽ đè dòng trước

    max_day = get_days_in_month(month)
    days = []

    for d in range(1, max_day + 1):
        if d in day_map:
            r = day_map[d]
            days.append(
                {
                    "day": d,  # FE đang parse lại thành dayNumber, nên số là ok nhất
                    "oee": float(r.get("OEERatio") or 0.0),
                    "ok_ratio": float(r.get("OKProductRatio") or 0.0),
                    "output_ratio": float(r.get("OutputRatio") or 0.0),
                    "activity_ratio": float(r.get("ActivityRatio") or 0.0),
                }
            )
        else:
            # ngày không có dữ liệu => vẫn trả, nhưng = 0
            days.append(
                {
                    "day": d,
                    "oee": 0.0,
                    "ok_ratio": 0.0,
                    "output_ratio": 0.0,
                    "activity_ratio": 0.0,
                }
            )

    return jsonify(
        {
            "machine_id": machine_id,
            "month": month,
            "data_type": data_type or None,
            "days": days,
        }
    )
@app.route("/api/machines/<int:machine_id>/month")
def get_machine_month_time(machine_id):
    try:
        month = int(request.args.get("month"))
    except (TypeError, ValueError):
        return jsonify({"error": "Missing or invalid month param"}), 400

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    # Lấy TỪNG DÒNG theo ngày (không GROUP / SUM trong SQL)
    cursor.execute(
        """
        SELECT
            Days,
            Operation,
            SmallStop,
            Fault,
            Break,
            Maintenance,
            Eat,
            Waiting,
            MachineryEdit,
            ChangeProductCode,
            Glue_CleaningPaper,
            Others
        FROM sdvn.dayvalues
        WHERE YEAR(Days)= %s AND MachineID = %s
          AND MONTH(Days) = %s
        ORDER BY Days
        """,
        (nam,machine_id, month),
    )

    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    # map theo ngày => categories
    day_map = {}

    # totals tháng (để FE hiển thị tổng, nếu cần)
    monthly_totals = {
        "Operation": 0.0,
        "SmallStop": 0.0,
        "Fault": 0.0,
        "Break": 0.0,
        "Maintenance": 0.0,
        "Eat": 0.0,
        "Waiting": 0.0,
        "MachineryEdit": 0.0,
        "ChangeProductCode": 0.0,
        "Glue_CleaningPaper": 0.0,
        "Others": 0.0,
    }

    for row in rows:
        day_raw = row["Days"]
        if hasattr(day_raw, "day"):
            dnum = day_raw.day
        else:
            try:
                dnum = int(str(day_raw)[-2:])
            except ValueError:
                continue

        categories = {
            "Operation": float(row.get("Operation") or 0.0),
            "SmallStop": float(row.get("SmallStop") or 0.0),
            "Fault": float(row.get("Fault") or 0.0),
            "Break": float(row.get("Break") or 0.0),
            "Maintenance": float(row.get("Maintenance") or 0.0),
            "Eat": float(row.get("Eat") or 0.0),
            "Waiting": float(row.get("Waiting") or 0.0),
            "MachineryEdit": float(row.get("MachineryEdit") or 0.0),
            "ChangeProductCode": float(row.get("ChangeProductCode") or 0.0),
            "Glue_CleaningPaper": float(row.get("Glue_CleaningPaper") or 0.0),
            "Others": float(row.get("Others") or 0.0),
        }

        day_map[dnum] = categories

    max_day = get_days_in_month(month)
    days = []

    for d in range(1, max_day + 1):
        if d in day_map:
            categories = day_map[d]
        else:
            categories = {
                "Operation": 0.0,
                "SmallStop": 0.0,
                "Fault": 0.0,
                "Break": 0.0,
                "Maintenance": 0.0,
                "Eat": 0.0,
                "Waiting": 0.0,
                "MachineryEdit": 0.0,
                "ChangeProductCode": 0.0,
                "Glue_CleaningPaper": 0.0,
                "Others": 0.0,
            }

        # cộng dồn totals
        for k in monthly_totals:
            monthly_totals[k] += categories[k]

        days.append(
            {
                "day": d,
                "categories": categories,
            }
        )

    result = {
        "machine_id": machine_id,
        "month": month,
        "days": days,
        "monthly_totals": {k: round(v, 2) for k, v in monthly_totals.items()},
    }

    return jsonify(result)
@app.route("/api/machines/<int:machine_id>/month-export", methods=["GET"])
def export_machine_month_excel(machine_id):
    try:
        month = int(request.args.get("month"))
    except (TypeError, ValueError):
        return jsonify({"error": "Missing or invalid month param"}), 400

    data_type = request.args.get("data", "ALL")

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    # --- 1. LẤY TÊN MÁY THEO ID ---
    # Nếu bạn có bảng machines(MachineID, MachineName) thì sửa query này lại cho đúng
    cursor.execute(
        "SELECT MachineName FROM machine WHERE MachineID = %s",
        (machine_id,),
    )
    mrow = cursor.fetchone()
    machine_name = (
        mrow["MachineName"] if mrow and mrow.get("MachineName") else f"Machine_{machine_id}"
    )

    # --- 2. LẤY DỮ LIỆU THÁNG TỪ dayvalues ---
    cursor.execute(
        """
        SELECT
            Days,
            OEERatio,
            OKProductRatio,
            OutputRatio,
            ActivityRatio,
            Operation,
            SmallStop,
            Fault,
            Break,
            Maintenance,
            Eat,
            Waiting,
            MachineryEdit,
            ChangeProductCode,
            Glue_CleaningPaper,
            Others
        FROM sdvn.dayvalues
        WHERE MachineID = %s
          AND MONTH(Days) = %s
        ORDER BY Days
        """,
        (machine_id, month),
    )
    rows = cursor.fetchall()

    cursor.close()
    conn.close()

    # --- 3. TẠO EXCEL ---
    wb = Workbook()
    ws = wb.active
    ws.title = machine_name

    # Dòng thông tin chung
    ws.append(
        [
            f"Machine: {machine_name}",
            f"Month: {month}",
            f"Data filter: {data_type}",
        ]
    )
    ws.append([])

    # Header cột
    headers = [
        "Date",
        # Ratio
        "OEERatio",
        "OKProductRatio",
        "OutputRatio",
        "ActivityRatio",
        # Time (giờ)
        "Operation",
        "SmallStop",
        "Fault",
        "Break",
        "Maintenance",
        "Eat",
        "Waiting",
        "MachineryEdit",
        "ChangeProductCode",
        "Glue_CleaningPaper",
        "Others",
        # Time (%)
        "OperationPct",
        "SmallStopPct",
        "FaultPct",
        "BreakPct",
        "MaintenancePct",
        "EatPct",
        "WaitingPct",
        "MachineryEditPct",
        "ChangeProductCodePct",
        "Glue_CleaningPaperPct",
        "OthersPct",
    ]
    ws.append(headers)

    # Ghi từng ngày
    for row in rows:
        # Date
        day_raw = row["Days"]
        day_str = (
            day_raw.strftime("%Y-%m-%d")
            if hasattr(day_raw, "strftime")
            else str(day_raw)
        )

        # Ratio
        oee = float(row.get("OEERatio") or 0.0)
        okr = float(row.get("OKProductRatio") or 0.0)
        out = float(row.get("OutputRatio") or 0.0)
        act = float(row.get("ActivityRatio") or 0.0)

        # Time (giờ)
        op = float(row.get("Operation") or 0.0)
        ss = float(row.get("SmallStop") or 0.0)
        flt = float(row.get("Fault") or 0.0)
        brk = float(row.get("Break") or 0.0)
        mt = float(row.get("Maintenance") or 0.0)
        eat = float(row.get("Eat") or 0.0)
        wait = float(row.get("Waiting") or 0.0)
        me = float(row.get("MachineryEdit") or 0.0)
        cpc = float(row.get("ChangeProductCode") or 0.0)
        gcp = float(row.get("Glue_CleaningPaper") or 0.0)
        oth = float(row.get("Others") or 0.0)

        total_time = (
            op
            + ss
            + flt
            + brk
            + mt
            + eat
            + wait
            + me
            + cpc
            + gcp
            + oth
        )

        def pct(val: float) -> float:
            if total_time <= 0:
                return 0.0
            return round((val * 100.0) / total_time, 2)

        ws.append(
            [
                day_str,
                # Ratio
                oee,
                okr,
                out,
                act,
                # Time (giờ)
                op,
                ss,
                flt,
                brk,
                mt,
                eat,
                wait,
                me,
                cpc,
                gcp,
                oth,
                # Time (%)
                pct(op),
                pct(ss),
                pct(flt),
                pct(brk),
                pct(mt),
                pct(eat),
                pct(wait),
                pct(me),
                pct(cpc),
                pct(gcp),
                pct(oth),
            ]
        )

    # --- 4. KẺ BẢNG (BORDER) CHO TẤT CẢ Ô ---
    thin = Side(style="thin")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    max_col = ws.max_column

    for row in ws.iter_rows(
        min_row=1, max_row=max_row, min_col=1, max_col=max_col
    ):
        for cell in row:
            cell.border = thin_border

    # --- 5. LƯU RA BUFFER ---
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # --- 6. TÊN FILE: tenmay_thang.xlsx (VD: LINE_01_09.xlsx) ---
    #safe_name = "".join(
      #  ch if ch.isalnum() or ch == " " else "_" for ch in machine_name
   # )
   # safe_name = safe_name.replace(" ", "_")
    filename = f"{machine_name}_{month:02d}.xlsx"

    try:
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except TypeError:
        # fallback nếu Flask cũ
        return send_file(
            output,
            as_attachment=True,
            attachment_filename=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
from flask import request, jsonify

@app.route("/api/machines/<int:machine_id>/year-ratio", methods=["GET"])
def get_machine_year_ratio(machine_id):
    """
    Ratio theo NĂM, luôn trả đủ 12 tháng.
    Tháng không có dữ liệu => ratio = 0
    """
    try:
        year = int(request.args.get("year"))
    except (TypeError, ValueError):
        return jsonify({"error": "Missing or invalid year param"}), 400

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute(
        """
        SELECT
            MONTH(Days) AS m,
            AVG(OEERatio)       AS avg_oee,
            AVG(OKProductRatio) AS avg_ok,
            AVG(OutputRatio)    AS avg_output,
            AVG(ActivityRatio)  AS avg_activity
        FROM sdvn.dayvalues
        WHERE MachineID = %s
          AND YEAR(Days) = %s
        GROUP BY MONTH(Days)
        ORDER BY m
        """,
        (machine_id, year),
    )
    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    # Chuyển thành dict theo tháng
    month_map = {int(r["m"]): r for r in rows}

    months = []
    for m in range(1, 13):  # luôn trả đủ 12 tháng
        if m in month_map:
            r = month_map[m]
            months.append(
                {
                    "month": m,
                    "oee": float(r.get("avg_oee") or 0.0),
                    "ok_ratio": float(r.get("avg_ok") or 0.0),
                    "output_ratio": float(r.get("avg_output") or 0.0),
                    "activity_ratio": float(r.get("avg_activity") or 0.0),
                }
            )
        else:
            months.append(
                {
                    "month": m,
                    "oee": 0,
                    "ok_ratio": 0,
                    "output_ratio": 0,
                    "activity_ratio": 0,
                }
            )

    return jsonify({"months": months})
@app.route("/api/machines/<int:machine_id>/year", methods=["GET"])
def get_machine_year_time(machine_id):
    """
    Thời gian theo NĂM, luôn trả đủ 12 tháng.
    Tháng không có dữ liệu => tất cả các field = 0
    """
    try:
        year = int(request.args.get("year"))
    except (TypeError, ValueError):
        return jsonify({"error": "Missing or invalid year param"}), 400

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute(
        """
        SELECT
            MONTH(Days) AS m,
            SUM(Operation)          AS op,
            SUM(SmallStop)          AS ss,
            SUM(Fault)              AS flt,
            SUM(`Break`)            AS brk,
            SUM(Maintenance)        AS mt,
            SUM(Eat)                AS eat,
            SUM(Waiting)            AS w,
            SUM(MachineryEdit)      AS me,
            SUM(ChangeProductCode)  AS cpc,
            SUM(Glue_CleaningPaper) AS gcp,
            SUM(Others)             AS oth
        FROM sdvn.dayvalues
        WHERE MachineID = %s
          AND YEAR(Days) = %s
        GROUP BY MONTH(Days)
        ORDER BY m
        """,
        (machine_id, year),
    )
    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    # Dict theo tháng
    month_map = {int(r["m"]): r for r in rows}

    months = []
    for m in range(1, 13):  # luôn trả 1..12
        if m in month_map:
            r = month_map[m]
            months.append(
                {
                    "month": m,
                    "categories": {
                        "Operation": float(r.get("op") or 0.0),
                        "SmallStop": float(r.get("ss") or 0.0),
                        "Fault": float(r.get("flt") or 0.0),
                        "Break": float(r.get("brk") or 0.0),
                        "Maintenance": float(r.get("mt") or 0.0),
                        "Eat": float(r.get("eat") or 0.0),
                        "Waiting": float(r.get("w") or 0.0),
                        "MachineryEdit": float(r.get("me") or 0.0),
                        "ChangeProductCode": float(r.get("cpc") or 0.0),
                        "Glue_CleaningPaper": float(r.get("gcp") or 0.0),
                        "Others": float(r.get("oth") or 0.0),
                    },
                }
            )
        else:
            months.append(
                {
                    "month": m,
                    "categories": {
                        "Operation": 0,
                        "SmallStop": 0,
                        "Fault": 0,
                        "Break": 0,
                        "Maintenance": 0,
                        "Eat": 0,
                        "Waiting": 0,
                        "MachineryEdit": 0,
                        "ChangeProductCode": 0,
                        "Glue_CleaningPaper": 0,
                        "Others": 0,
                    },
                }
            )

    return jsonify({"months": months})
@app.route("/api/machines/<int:machine_id>/year-export", methods=["GET"])
def export_machine_year_excel(machine_id):
    """
    Xuất Excel (.xlsx) dữ liệu NĂM cho 1 máy – 1 sheet, 12 dòng (tháng 1..12)

    Mỗi dòng:
    - Month
    - OEERatio, OKProductRatio, OutputRatio, ActivityRatio (TB theo tháng)
    - Operation..Others (tổng giờ theo tháng)
    - OperationPct..OthersPct (tỷ lệ %, làm tròn 2 số)
    """
    # --- Lấy YEAR ---
    try:
        year = int(request.args.get("year"))
    except (TypeError, ValueError):
        return jsonify({"error": "Missing or invalid year param"}), 400

    data_type = request.args.get("data", "ALL")  # để ghi chú trong header

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    # --- 1. LẤY TÊN MÁY THEO ID (productionline) ---
    cursor.execute(
        "SELECT MachineName FROM machine WHERE MachineID = %s",
        (machine_id,),
    )
    mrow = cursor.fetchone()
    machine_name = (
        mrow["MachineName"] if mrow and mrow.get("MachineName") else f"Machine_{machine_id}"
    )

    # --- 2. LẤY DỮ LIỆU NĂM (GỘP THEO THÁNG) ---
    cursor.execute(
        """
        SELECT
            MONTH(Days)           AS m,
            AVG(OEERatio)         AS avg_oee,
            AVG(OKProductRatio)   AS avg_ok,
            AVG(OutputRatio)      AS avg_output,
            AVG(ActivityRatio)    AS avg_activity,
            SUM(Operation)        AS sum_op,
            SUM(SmallStop)        AS sum_small,
            SUM(Fault)            AS sum_fault,
            SUM(`Break`)          AS sum_break,
            SUM(Maintenance)      AS sum_maint,
            SUM(Eat)              AS sum_eat,
            SUM(Waiting)          AS sum_wait,
            SUM(MachineryEdit)    AS sum_me,
            SUM(ChangeProductCode)  AS sum_cpc,
            SUM(Glue_CleaningPaper) AS sum_gcp,
            SUM(Others)             AS sum_oth
        FROM sdvn.dayvalues
        WHERE MachineID = %s
          AND YEAR(Days) = %s
        GROUP BY MONTH(Days)
        ORDER BY m
        """,
        (machine_id, year),
    )
    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    # map theo tháng
    month_map = {int(r["m"]): r for r in rows}

    # --- 3. TẠO EXCEL ---
    wb = Workbook()
    ws = wb.active
    ws.title = machine_name

    # Header thông tin chung
    ws.append(
        [
            f"MachineName: {machine_name}",
            f"Year: {year}",
            f"Data filter: {data_type}",
        ]
    )
    ws.append([])

    # Header cột
    headers = [
        "Month",
        # Ratio
        "OEERatio",
        "OKProductRatio",
        "OutputRatio",
        "ActivityRatio",
        # Time (giờ)
        "Operation",
        "SmallStop",
        "Fault",
        "Break",
        "Maintenance",
        "Eat",
        "Waiting",
        "MachineryEdit",
        "ChangeProductCode",
        "Glue_CleaningPaper",
        "Others",
        # Time (%)
        "OperationPct",
        "SmallStopPct",
        "FaultPct",
        "BreakPct",
        "MaintenancePct",
        "EatPct",
        "WaitingPct",
        "MachineryEditPct",
        "ChangeProductCodePct",
        "Glue_CleaningPaperPct",
        "OthersPct",
    ]
    ws.append(headers)

    # helper tính %
    def pct_part(val, total):
        if not total or total <= 0:
            return 0.0
        return round((val * 100.0) / total, 2)

    # Ghi đủ 12 tháng
    for m in range(1, 13):
        r = month_map.get(m)

        if r:
            oee = float(r.get("avg_oee") or 0.0)
            okr = float(r.get("avg_ok") or 0.0)
            out = float(r.get("avg_output") or 0.0)
            act = float(r.get("avg_activity") or 0.0)

            op   = float(r.get("sum_op") or 0.0)
            ss   = float(r.get("sum_small") or 0.0)
            flt  = float(r.get("sum_fault") or 0.0)
            brk  = float(r.get("sum_break") or 0.0)
            mt   = float(r.get("sum_maint") or 0.0)
            eat  = float(r.get("sum_eat") or 0.0)
            wait = float(r.get("sum_wait") or 0.0)
            me   = float(r.get("sum_me") or 0.0)
            cpc  = float(r.get("sum_cpc") or 0.0)
            gcp  = float(r.get("sum_gcp") or 0.0)
            oth  = float(r.get("sum_oth") or 0.0)
        else:
            oee = okr = out = act = 0.0
            op = ss = flt = brk = mt = eat = wait = me = cpc = gcp = oth = 0.0

        total_time = (
            op
            + ss
            + flt
            + brk
            + mt
            + eat
            + wait
            + me
            + cpc
            + gcp
            + oth
        )

        ws.append(
            [
                m,          # Month
                # Ratio
                oee,
                okr,
                out,
                act,
                # Time
                op,
                ss,
                flt,
                brk,
                mt,
                eat,
                wait,
                me,
                cpc,
                gcp,
                oth,
                # Time %
                pct_part(op, total_time),
                pct_part(ss, total_time),
                pct_part(flt, total_time),
                pct_part(brk, total_time),
                pct_part(mt, total_time),
                pct_part(eat, total_time),
                pct_part(wait, total_time),
                pct_part(me, total_time),
                pct_part(cpc, total_time),
                pct_part(gcp, total_time),
                pct_part(oth, total_time),
            ]
        )

    # --- 4. KẺ BẢNG (BORDER) ---
    thin = Side(style="thin")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    max_col = ws.max_column
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border

    # --- 5. GHI RA BUFFER & TÊN FILE ---
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = "".join(
        ch if ch.isalnum() or ch == " " else "_" for ch in machine_name
    )
    safe_name = safe_name.replace(" ", "_")
    filename = f"{safe_name}_nam_{year}.xlsx"

    try:
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except TypeError:
        # nếu Flask cũ
        return send_file(
            output,
            as_attachment=True,
            attachment_filename=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
@app.route("/api/line-kpi", methods=["GET"])
def get_line_kpi():
    line = request.args.get("line")            # Line550B, Line400B...
    month = request.args.get("month")          # "7"
    year = request.args.get("year")            # "2025"
    data_type = request.args.get("data", "all")  # 👈 all / oee / ok / output / activity (từ FE)

    now = datetime.now()
    month = int(month) if month else now.month
    year = int(year) if year else now.year

    if not line:
        return jsonify({"error": "Missing 'line' parameter"}), 400

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        query = """
            SELECT 
                pl.LineName,
                dv.Days,
                SUM(dv.OEERatio)       AS total_OEERatio,
                SUM(dv.OKProductRatio) AS total_OKProductRatio,
                SUM(dv.OutputRatio)    AS total_OutputRatio,
                SUM(dv.ActivityRatio)  AS total_ActivityRatio
            FROM dayvalues dv
            JOIN machine m         ON dv.MachineID = m.MachineID
            JOIN productionline pl ON m.LineID = pl.LineID
            WHERE MONTH(dv.Days) = %s
              AND YEAR(dv.Days) = %s
              AND pl.LineName = %s
            GROUP BY pl.LineName, dv.Days
            ORDER BY dv.Days
        """

        cursor.execute(query, (month, year, line))
        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        chart_data = []
        for r in rows:
            d = r["Days"]
            if isinstance(d, datetime):
                day_num = d.day
            else:
                day_num = int(str(d).split("-")[2])

            chart_data.append({
                "LineName": r["LineName"],
                "day": day_num,
                "oee": float(r["total_OEERatio"] or 0),
                "ok": float(r["total_OKProductRatio"] or 0),
                "output": float(r["total_OutputRatio"] or 0),
                "activity": float(r["total_ActivityRatio"] or 0),
                "data_type": data_type,   # 👈 có dùng param data (lưu lại, sau cần phân tích/log)
            })

        return jsonify(chart_data), 200

    except Exception as e:
        print("Unknown error in /api/line-kpi:", e)
        return jsonify({"error": "Server error"}), 500
@app.route("/api/export-kpi", methods=["GET"])
def export_kpi():
    # Lấy param từ FE
    month = request.args.get("month", type=int)
    year = request.args.get("year", type=int)
    data_type = request.args.get("data", default="all")  # all / oee / ok / output / activity (để dành)

    now = datetime.now()
    if not month:
        month = now.month
    if not year:
        year = now.year

    # Border mảnh cho bảng
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # Lấy KPI cho TẤT CẢ line trong tháng/năm
        query = """
            SELECT 
                pl.LineName,
                dv.Days,
                SUM(dv.OEERatio)       AS total_OEERatio,
                SUM(dv.OKProductRatio) AS total_OKProductRatio,
                SUM(dv.OutputRatio)    AS total_OutputRatio,
                SUM(dv.ActivityRatio)  AS total_ActivityRatio
            FROM dayvalues dv
            JOIN machine m         ON dv.MachineID = m.MachineID
            JOIN productionline pl ON m.LineID = pl.LineID
            WHERE MONTH(dv.Days) = %s
              AND YEAR(dv.Days) = %s
            GROUP BY pl.LineName, dv.Days
            ORDER BY pl.LineName, dv.Days
        """
        cursor.execute(query, (month, year))
        rows = cursor.fetchall()

        cursor.close()
        conn.close()

        # Gom data theo line -> data_by_line[line_name][day] = metrics
        data_by_line = {}
        for r in rows:
            line_name = r["LineName"]
            d = r["Days"]
            if isinstance(d, datetime):
                day_num = d.day
            else:
                day_num = int(str(d).split("-")[2])

            if line_name not in data_by_line:
                data_by_line[line_name] = {}

            data_by_line[line_name][day_num] = {
                "oee": float(r["total_OEERatio"] or 0),
                "ok": float(r["total_OKProductRatio"] or 0),
                "output": float(r["total_OutputRatio"] or 0),
                "activity": float(r["total_ActivityRatio"] or 0),
            }

        # Tạo workbook Excel
        wb = Workbook()
        # Xoá sheet mặc định
        default_ws = wb.active
        wb.remove(default_ws)

        # Số ngày trong tháng (tự xử lý 28/29/30/31)
        days_in_month = calendar.monthrange(year, month)[1]

        if not data_by_line:
            # Không có line nào => vẫn tạo 1 sheet NoData
            ws = wb.create_sheet(title="NoData")
            ws.append(["Day", "OEERatio", "OKProductRatio", "OutputRatio", "ActivityRatio"])

            for day in range(1, days_in_month + 1):
                ws.append([day, 0, 0, 0, 0])

            # Kẻ border
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5):
                for cell in row:
                    cell.border = thin_border

        else:
            # Mỗi line 1 sheet
            for line_name, day_map in data_by_line.items():
                sheet_name = (line_name or "Line")[:31]  # Sheet name max 31 ký tự
                ws = wb.create_sheet(title=sheet_name)

                # Header
                ws.append(["Day", "OEERatio", "OKProductRatio", "OutputRatio", "ActivityRatio"])

                # Duyệt tất cả ngày trong tháng, fill 0 nếu không có
                for day in range(1, days_in_month + 1):
                    metrics = day_map.get(day, {})
                    oee = metrics.get("oee", 0)
                    ok = metrics.get("ok", 0)
                    output = metrics.get("output", 0)
                    activity = metrics.get("activity", 0)
                    ws.append([day, oee, ok, output, activity])

                # Kẻ border cho toàn sheet
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5):
                    for cell in row:
                        cell.border = thin_border

        # Xuất ra memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        file_name = f"OverView_T_{month}_{year}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=file_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


    except Exception as e:
        print("Unknown error in /api/export-kpi:", e)
        return jsonify({"error": "Server error"}), 500
    return jsonify(result)
@app.route("/api/day-plans", methods=["GET"])
def get_day_plans():
    idline = request.args.get("idline", type=int)
    idmachine = request.args.get("idmachine", type=int)
    date = request.args.get("date")

    if not idline or not date:
        return jsonify({"error": "Missing params"}), 400

    db = get_connection()
    cursor = db.cursor()

    # === STEP 1: kiểm tra có dữ liệu chưa ===
    cursor.execute("""
        SELECT dv.idplan_production
        FROM plan_production dv 
        JOIN machine m ON dv.MachineID = m.MachineID
        WHERE m.LineID = %s
          AND dv.Days = %s
    """, (idline, date))

    existing = cursor.fetchall()

    # === STEP 2: nếu chưa có → INSERT tất cả máy theo line ===
    if len(existing) == 0:
        cursor.execute("""
            SELECT MachineID, CycleTime 
            FROM machine 
            WHERE LineID = %s
        """, (idline,))
        machines = cursor.fetchall()

        for (mid, cycletime) in machines:
            cursor.execute("""
                INSERT INTO plan_production
                (MachineID, Days, DayPlan, Target_Product,
                 StartTime_1, EndTime_1, StartTime_2, EndTime_2)
                VALUES (%s, %s, 16, 0,
                        %s, %s, %s, %s)
            """, (
                mid, date,
                f"{date} 06:00:00",
                f"{date} 14:00:00",
                f"{date} 14:00:00",
                f"{date} 22:00:00"
            ))

        db.commit()

    # === STEP 3: Lấy lại dữ liệu sau khi insert (nếu có) ===
    sql = """
        SELECT 
            pl.LineName,
            m.MachineName,
            dv.Days,
            dv.DayPlan,
            dv.Target_Product,
            m.CycleTime,
            dv.StartTime_1,
            dv.EndTime_1,
            dv.StartTime_2,
            dv.EndTime_2,
            dv.idplan_production
        FROM plan_production dv
        JOIN machine m ON dv.MachineID = m.MachineID
        JOIN productionline pl ON m.LineID = pl.LineID
        WHERE m.LineID = %s
          AND dv.Days = %s
    """

    params = [idline, date]

    if idmachine:  # nếu FE chọn lọc 1 máy
        sql += " AND m.MachineID = %s"
        params.append(idmachine)

    sql += " ORDER BY m.MachineID"

    cursor.execute(sql, params)
    rows = cursor.fetchall()

    return jsonify(format_rows(rows))
@app.route("/api/day-plans/bulk-update", methods=["PUT"])
def bulk_update_day_plans():
    db = get_connection()
    cursor = db.cursor()

    plans = request.get_json() or []

    def parse_dt(s):
        from datetime import datetime
        if not s:
            return None
        try:
            if len(s) == 16:
                return datetime.strptime(s, "%Y-%m-%dT%H:%M")
            return datetime.fromisoformat(s)
        except Exception:
            return None

    for p in plans:
        if not isinstance(p, dict):
            continue

        plan_id = p.get("id")
        if not plan_id:
            continue

        s1 = parse_dt(p.get("startShift1"))
        e1 = parse_dt(p.get("endShift1"))
        s2 = parse_dt(p.get("startShift2"))
        e2 = parse_dt(p.get("endShift2"))

        new_cycle_time = p.get("cycleTime")  # FE nhập

        # 1. Lấy MachineID + CycleTime hiện tại
        cursor.execute("""
            SELECT dv.MachineID, m.CycleTime
            FROM plan_production dv
            JOIN machine m ON dv.MachineID = m.MachineID
            WHERE dv.idplan_production = %s
        """, (plan_id,))
        row = cursor.fetchone()
        if not row:
            continue

        machine_id, current_ct = row

        # 2. Quyết định dùng cycleTime nào để TÍNH
        cycle_time = None
        try:
            if new_cycle_time is not None and new_cycle_time != "":
                cycle_time = float(new_cycle_time)   # ƯU TIÊN CYCLE MỚI
            else:
                cycle_time = float(current_ct) if current_ct is not None else 0
        except Exception:
            cycle_time = float(current_ct) if current_ct is not None else 0

        # 3. Tính DayPlan = (End1-Start1)+(End2-Start2) (giờ, kiểu int)
        diff1 = (e1 - s1).total_seconds()/3600 if (s1 and e1) else 0
        diff2 = (e2 - s2).total_seconds()/3600 if (s2 and e2) else 0
        day_plan = int(round(diff1 + diff2))

        # 4. Target_Product = DayPlan * 3600 / CycleTime (dùng cycle_time MỚI)
        target_product = int(day_plan * 3600 / cycle_time) if cycle_time and cycle_time > 0 else 0

        # 5. UPDATE bảng plan_production
        cursor.execute("""
            UPDATE plan_production
            SET
                DayPlan = %s,
                Target_Product = %s,
                StartTime_1 = %s,
                EndTime_1 = %s,
                StartTime_2 = %s,
                EndTime_2 = %s
            WHERE idplan_production = %s
        """, (
            day_plan,
            target_product,
            s1.strftime("%Y-%m-%d %H:%M:%S") if s1 else None,
            e1.strftime("%Y-%m-%d %H:%M:%S") if e1 else None,
            s2.strftime("%Y-%m-%d %H:%M:%S") if s2 else None,
            e2.strftime("%Y-%m-%d %H:%M:%S") if e2 else None,
            plan_id,
        ))

        # 6. UPDATE CycleTime vào bảng machine (nếu FE có sửa)
        if new_cycle_time is not None and new_cycle_time != "":
            try:
                ct_val = float(new_cycle_time)
                cursor.execute("""
                    UPDATE machine
                    SET CycleTime = %s
                    WHERE MachineID = %s
                """, (ct_val, machine_id))
            except Exception:
                pass

    db.commit()
    return jsonify({"status": "ok", "updated": len(plans)})
@app.route("/api/month-plans", methods=["GET"])
def get_month_plans():
    idline = request.args.get("idline", type=int)
    idmachine = request.args.get("idmachine", type=int)
    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)

    if not idline or not year or not month:
        return jsonify({"error": "Missing params"}), 400

    db = get_connection()
    cursor = db.cursor()

    # Số ngày trong tháng
    last_day = calendar.monthrange(year, month)[1]
    all_days = [f"{year}-{month:02d}-{d:02d}" for d in range(1, last_day + 1)]

    # Lấy danh sách máy
    if idmachine:
        # FE gửi "All" thì coi như None
        try:
            machine_id_int = int(idmachine)
        except:
            machine_id_int = None
    else:
        machine_id_int = None

    if machine_id_int:
        cursor.execute("""
            SELECT MachineID, CycleTime 
            FROM machine
            WHERE MachineID = %s
        """, (machine_id_int,))
    else:
        cursor.execute("""
            SELECT MachineID, CycleTime
            FROM machine
            WHERE LineID = %s
        """, (idline,))

    machines = cursor.fetchall()  # (MachineID, CycleTime)

    # Với mỗi máy: nếu thiếu ngày nào trong tháng → insert
    for (mid, cycletime) in machines:
        cursor.execute("""
            SELECT Days
            FROM plan_production
            WHERE MachineID = %s
              AND YEAR(Days) = %s
              AND MONTH(Days) = %s
        """, (mid, year, month))

        existing = {row[0].strftime("%Y-%m-%d") for row in cursor.fetchall()}
        missing_days = [d for d in all_days if d not in existing]

        for day in missing_days:
            cursor.execute("""
                INSERT INTO plan_production
                (MachineID, Days, DayPlan, Target_Product,
                 StartTime_1, EndTime_1, StartTime_2, EndTime_2)
                VALUES (%s, %s, 0, 0,
                        %s, %s, %s, %s)
            """, (
                mid, day,
                f"{day} 06:00:00",
                f"{day} 14:00:00",
                f"{day} 14:00:00",
                f"{day} 22:00:00",
            ))

    db.commit()

    # Lấy dữ liệu trả về cho FE
    sql = """
        SELECT 
            pl.LineName,
            m.MachineName,
            dv.Days,
            dv.DayPlan,
            dv.Target_Product,
            m.CycleTime,
            dv.StartTime_1,
            dv.EndTime_1,
            dv.StartTime_2,
            dv.EndTime_2,
            dv.idplan_production
        FROM plan_production dv
        JOIN machine m ON dv.MachineID = m.MachineID
        JOIN productionline pl ON m.LineID = pl.LineID
        WHERE m.LineID = %s
          AND YEAR(dv.Days) = %s
          AND MONTH(dv.Days) = %s
    """

    params = [idline, year, month]

    if machine_id_int:
        sql += " AND dv.MachineID = %s"
        params.append(machine_id_int)

    sql += " ORDER BY m.MachineID, dv.Days"

    cursor.execute(sql, params)
    rows = cursor.fetchall()

    return jsonify(format_rows(rows))
@app.route("/api/month-plans/bulk-update", methods=["PUT"])
def bulk_update_month_plans():
    db = get_connection()
    cursor = db.cursor()

    plans = request.get_json() or []
    if not isinstance(plans, list):
        return jsonify({"error": "Body must be an array"}), 400

    # --- HÀM HỖ TRỢ ---
    def parse_dt(s):
        if not s:
            return None
        try:
            if len(s) == 16:  # yyyy-MM-ddTHH:mm
                return datetime.strptime(s, "%Y-%m-%dT%H:%M")
            return datetime.fromisoformat(s)
        except:
            return None

    # Map máy có CT mới → only update nếu có thay đổi
    machine_new_cycle = {}

    # ================================
    # 1) Quét các dòng để xem cycle time có thay đổi không
    # ================================
    for p in plans:
        if not isinstance(p, dict):
            continue

        plan_id = p.get("id")
        if not plan_id:
            continue

        new_ct_val = p.get("cycleTime")

        # lấy machineID + cycleTime hiện tại trong DB
        cursor.execute("""
            SELECT dv.MachineID, m.CycleTime
            FROM plan_production dv
            JOIN machine m ON dv.MachineID = m.MachineID
            WHERE dv.idplan_production = %s
        """, (plan_id,))
        row = cursor.fetchone()

        if not row:
            continue

        machine_id, old_cycle = row

        # Nếu FE không gửi cycle → bỏ qua
        if new_ct_val is None or new_ct_val == "":
            continue

        try:
            new_ct_int = int(new_ct_val)  # ép kiểu int
        except:
            continue

        # Nếu cycleTime KHÁC DB → cần update
        if old_cycle is None or int(old_cycle) != new_ct_int:
            machine_new_cycle[machine_id] = new_ct_int

    # ================================
    # 2) Update cycle time chỉ những máy được sửa
    # ================================
    for mid, ct in machine_new_cycle.items():
        print(">>> UPDATE MACHINE: MachineID =", mid, ", Cycle =", ct)
        cursor.execute("""
            UPDATE machine SET CycleTime = %s WHERE MachineID = %s
        """, (ct, mid))

    db.commit()   # commit riêng cycle time cho chắc chắn

    # ================================
    # 3) Lấy lại cycle time mới từ DB
    # ================================
    cycle_map = {}
    for mid in machine_new_cycle.keys():
        cursor.execute("SELECT CycleTime FROM machine WHERE MachineID = %s", (mid,))
        row = cursor.fetchone()
        if row:
            cycle_map[mid] = int(row[0])   # đảm bảo int

    # ================================
    # 4) Update lại từng plan
    # ================================
    for p in plans:

        plan_id = p.get("id")
        if not plan_id:
            continue

        # lấy machineID
        cursor.execute("""
            SELECT MachineID FROM plan_production WHERE idplan_production = %s
        """, (plan_id,))
        r = cursor.fetchone()
        if not r:
            continue

        machine_id = r[0]

        # lấy cycle time mới nhất
        if machine_id in cycle_map:
            cycle_time = cycle_map[machine_id]
        else:
            cursor.execute("SELECT CycleTime FROM machine WHERE MachineID = %s", (machine_id,))
            r2 = cursor.fetchone()
            cycle_time = int(r2[0]) if r2 else 0

        # parse time
        s1 = parse_dt(p.get("startShift1"))
        e1 = parse_dt(p.get("endShift1"))
        s2 = parse_dt(p.get("startShift2"))
        e2 = parse_dt(p.get("endShift2"))

        diff1 = (e1 - s1).total_seconds()/3600 if (s1 and e1) else 0
        diff2 = (e2 - s2).total_seconds()/3600 if (s2 and e2) else 0

        day_plan = int(diff1 + diff2)

        target_product = int(day_plan * 3600 / cycle_time) if cycle_time > 0 else 0

        cursor.execute("""
            UPDATE plan_production
            SET DayPlan = %s,
                Target_Product = %s,
                StartTime_1 = %s,
                EndTime_1 = %s,
                StartTime_2 = %s,
                EndTime_2 = %s
            WHERE idplan_production = %s
        """, (
            day_plan,
            target_product,
            s1.strftime("%Y-%m-%d %H:%M:%S") if s1 else None,
            e1.strftime("%Y-%m-%d %H:%M:%S") if e1 else None,
            s2.strftime("%Y-%m-%d %H:%M:%S") if s2 else None,
            e2.strftime("%Y-%m-%d %H:%M:%S") if e2 else None,
            plan_id,
        ))

    db.commit()
    return jsonify({"status": "ok"})
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)